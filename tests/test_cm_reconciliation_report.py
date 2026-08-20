# ===============================================================================
# Copyright 2026 ross
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ===============================================================================
"""Tests for the critical-minerals reconciliation report."""

import openpyxl
import pytest

from scripts.cm_reconciliation_report import build_report, classify, recommend

HEADERS = [
    "SAMPLE ",
    "Project",
    "Area",
    "Reference",
    "Date collected",
    "Date analyzed",
    "Laboratory",
    "latitude",
    "longitude",
    "Coordinate system",
    "SiO2",
    "TiO2",
    "Al2O3",
    "Fe2O3T",
    "MnO",
    "MgO",
    "CaO",
    "Na2O",
    "K2O",
    "P2O5",
    "LOI ",
    "F",
    "Total",
    "Au",
]
UNITS = [None] * 10 + ["%"] * 12 + ["ppb"]

# A clean row: majors sum to ~100, coordinates in New Mexico, real dates.
CLEAN = [
    "Gal1",
    "Gallinas",
    "Gallinas",
    "McLemore et al. (2025b)",
    "2019-08-28",
    "2020-04-27",
    "ALS",
    "34.19701153",
    "-105.74021543",
    "WGS84",
    60.0,
    0.5,
    16.0,
    5.0,
    0.1,
    2.0,
    5.0,
    3.0,
    5.0,
    0.2,
    3.2,
    0.5,
    100.5,
    12,
]


def _row(**overrides):
    row = list(CLEAN)
    for header, value in overrides.items():
        row[HEADERS.index(header)] = value
    return row


CHEMICAL_DATA_ROWS = [
    CLEAN,
    # F reported in ppm inside a column the workbook declares as %.
    _row(**{"SAMPLE ": "Gal2", "F": 27700}),
    # Total nowhere near the sum of the majors it is meant to add up.
    _row(**{"SAMPLE ": "Gal3", "Total": 60.0}),
    # Coordinates in Kansas, and a year-only collection date.
    _row(**{"SAMPLE ": "Gal4", "latitude": "39.5", "Date collected": "1993"}),
    # Text where an analyte value belongs, and no coordinates at all.
    _row(**{"SAMPLE ": "Gal5", "Au": "bd", "latitude": None, "longitude": None}),
    # A repeated sample name, so SAMPLE cannot be the key on its own.
    _row(**{"SAMPLE ": "Gal1", "Laboratory": "USGS"}),
    # Present in ChemicalData only, as the real workbook's tail rows are.
    _row(**{"SAMPLE ": "Gal6"}),
]
# GIS drops the last row and disagrees on Laboratory (blank) and Total (#VALUE!).
GIS_ROWS = [
    _row(Laboratory=None),
    _row(**{"SAMPLE ": "Gal2", "F": 27700}),
    _row(**{"SAMPLE ": "Gal3", "Total": "#VALUE!"}),
    _row(**{"SAMPLE ": "Gal4", "latitude": "39.5", "Date collected": "1993"}),
    _row(**{"SAMPLE ": "Gal5", "Au": "bd", "latitude": None, "longitude": None}),
    _row(**{"SAMPLE ": "Gal1", "Laboratory": "USGS"}),
]


@pytest.fixture()
def source_workbook(tmp_path):
    workbook = openpyxl.Workbook()
    chemical_data = workbook.active
    chemical_data.title = "ChemicalData"
    chemical_data.append(["Chemical analyses of samples"])
    chemical_data.append(HEADERS)
    chemical_data.append(["Units"] + UNITS[1:])
    for row in CHEMICAL_DATA_ROWS:
        chemical_data.append(row)

    gis = workbook.create_sheet("GIS")
    gis.append(HEADERS)
    for row in GIS_ROWS:
        gis.append(row)

    detection_limits = workbook.create_sheet("DetectionLimits")
    detection_limits.append(["Method C_ICPOES_MS-61"])
    detection_limits.append(["Element", "Lower Reporting Limit", "Unit"])
    detection_limits.append(["Au", 10, "ppb"])

    path = tmp_path / "McLemoreMasterChem_test.xlsx"
    workbook.save(path)
    return path


@pytest.fixture()
def report(source_workbook, tmp_path):
    output = tmp_path / "reconciliation.xlsx"
    summary = build_report(source_workbook, output)
    return summary, openpyxl.load_workbook(output)


def _records(worksheet):
    rows = list(worksheet.iter_rows(values_only=True))
    return [dict(zip(rows[0], row)) for row in rows[1:]]


@pytest.mark.parametrize(
    "chemical_data_value, gis_value, expected",
    [
        ("ALS", "ALS", None),
        (" ALS ", "ALS", None),
        (None, None, None),
        ("ALS", None, "only in ChemicalData"),
        (None, "ALS", "only in GIS"),
        ("100.5", "#VALUE!", "Excel error in GIS"),
        ("#VALUE!", "100.5", "Excel error in ChemicalData"),
        ("100.5", "60.0", "different number"),
        ("100.5", "100.50", "same number, different text"),
        ("ZuniMountains", "Zuni", "different text"),
    ],
)
def test_classify_names_the_kind_of_disagreement(
    chemical_data_value, gis_value, expected
):
    assert classify(chemical_data_value, gis_value) == expected


@pytest.mark.parametrize(
    "kinds, expected_fragment",
    [
        ({"only in GIS": 5}, "Take GIS"),
        ({"only in ChemicalData": 5}, "Take ChemicalData"),
        ({"only in GIS": 5, "only in ChemicalData": 2}, "Fill blanks"),
        ({"Excel error in GIS": 3}, "without the #VALUE!"),
        ({"different text": 1}, "per-row ruling"),
        # A conflict outranks fillable blanks: it cannot be resolved in bulk.
        ({"only in GIS": 50, "different number": 1}, "per-row ruling"),
    ],
)
def test_recommend_prefers_the_safest_reading(kinds, expected_fragment):
    assert expected_fragment in recommend("laboratory", kinds)


def test_report_has_a_sheet_for_every_decision(report):
    _, workbook = report

    assert workbook.sheetnames == [
        "README",
        "ColumnDecisions",
        "CellDifferences",
        "RowsOnlyInOneSheet",
        "IntegritySummary",
        "IntegrityDetail",
        "DetectionLimitSpread",
        "DuplicateSampleNames",
    ]


def test_summary_counts_the_drift(report):
    summary, _ = report

    assert summary["chemical_data_rows"] == 7
    assert summary["gis_rows"] == 6
    assert summary["aligned_rows"] == 6
    assert summary["rows_only_in_chemical_data"] == 1
    assert summary["rows_only_in_gis"] == 0
    assert summary["duplicate_sample_names"] == 1


def test_column_decisions_carry_a_blank_ruling_and_a_dropdown(report):
    _, workbook = report
    worksheet = workbook["ColumnDecisions"]
    records = _records(worksheet)

    laboratory = next(row for row in records if row["Mirror column"] == "laboratory")
    assert laboratory["Only in ChemicalData"] == 1
    assert "Take ChemicalData" in laboratory["Suggested starting point"]
    assert laboratory["DECISION"] is None

    total = next(row for row in records if row["Mirror column"] == "total_pct")
    assert total["Excel error"] == 1

    validations = worksheet.data_validations.dataValidation
    assert len(validations) == 1
    assert "ChemicalData" in validations[0].formula1


def test_cell_differences_show_both_values_side_by_side(report):
    _, workbook = report
    records = _records(workbook["CellDifferences"])

    laboratory = next(row for row in records if row["Mirror column"] == "laboratory")
    assert laboratory["ChemicalData value"] == "ALS"
    assert laboratory["GIS value"] is None
    assert laboratory["Kind of difference"] == "only in ChemicalData"
    # Row numbers point back into the delivered workbook.
    assert laboratory["ChemicalData row"] == 4
    assert laboratory["GIS row"] == 2


def test_rows_only_in_one_sheet_names_the_missing_sample(report):
    _, workbook = report
    records = _records(workbook["RowsOnlyInOneSheet"])

    assert [row["SAMPLE"] for row in records] == ["Gal6"]
    assert records[0]["Sheet"] == "ChemicalData"
    assert records[0]["Has coordinates"] == "yes"


def test_integrity_checks_flag_the_unit_contradiction(report):
    _, workbook = report
    records = _records(workbook["IntegrityDetail"])

    unit_issues = [
        row
        for row in records
        if row["Issue"] == "Value impossible for the declared unit"
    ]
    assert [row["SAMPLE"] for row in unit_issues] == ["Gal2"]
    assert unit_issues[0]["Column"] == "f_pct"
    assert unit_issues[0]["Value"] == "27700"
    assert "looks like ppm" in unit_issues[0]["Detail"]

    issues = {row["Issue"] for row in records}
    assert "Total disagrees with the sum of the major oxides" in issues
    assert "Latitude outside New Mexico" in issues
    assert "Non-numeric analyte value" in issues
    assert "No coordinates" in issues
    assert "Date is not a full date" in issues


def test_a_total_that_is_a_sum_is_not_treated_as_a_unit_error(report):
    """Total legitimately exceeds 100; only real measurements get the unit check."""
    _, workbook = report
    records = _records(workbook["IntegrityDetail"])

    assert not [
        row
        for row in records
        if row["Column"] == "total_pct"
        and row["Issue"] == "Value impossible for the declared unit"
    ]


def test_integrity_summary_explains_each_issue(report):
    _, workbook = report
    records = _records(workbook["IntegritySummary"])

    assert records == sorted(records, key=lambda row: -row["Rows affected"])
    for row in records:
        assert row["Why it matters"]
        assert row["RESOLUTION"] is None


def test_duplicate_sample_names_lists_the_source_rows(report):
    _, workbook = report
    records = _records(workbook["DuplicateSampleNames"])

    assert records[0]["SAMPLE"] == "Gal1"
    assert records[0]["Rows with this name"] == 2
    assert records[0]["Source rows"] == "4, 9"
