# ===============================================================================
# Copyright 2026 ross
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ===============================================================================
"""Tests for the CM_legacy critical-minerals staging mirror."""

from datetime import date, datetime
from decimal import Decimal

import openpyxl
import pytest
from sqlalchemy import delete, select

from db.cm_legacy import (
    ANALYTE_UNITS,
    CM_CHEMISTRY_SOURCE_SHEETS,
    QAQC_MISSING_COLUMNS,
    SOURCE_HEADER_BY_COLUMN,
    CM_ChemicalData,
    CM_DetectionLimits,
    CM_MineralSystems,
    CM_References,
    CM_WorkbookMetadata,
    CM_WorldComparisons,
    CM_WorldReferences,
)
from db.engine import session_ctx
from services.cm_legacy_mirror import (
    CMWorkbookError,
    cell_to_text,
    load_cm_workbook,
)

CHEM_HEADERS = [
    "SAMPLE ",
    "Project",
    "Area",
    "Date collected",
    "latitude",
    "longitude",
    "Coordinate system",
    "MapSymbol",
    "Depth/legnth (ft)",
    "paste pH",
    "SiO2",
    "Total",
    "Au",
    "As",
    "In",
    "Pd",
    "Pt",
]
CHEM_UNITS = [None] * 10 + ["%", "%", "ppb", "ppm", "ppm", "ppm", "ppm"]
CHEM_ROWS = [
    [
        "Gal1",
        "Gallinas",
        "Gallinas",
        datetime(2019, 8, 28),
        "34.19701153",
        "-105.74021543",
        "WGS84",
        "Tv",
        0,
        None,
        74.5,
        101.84,
        "<10",
        82,
        "<0.1",
        4,
        "<5",
    ],
    [
        "Gal2",
        "Gallinas",
        "ZuniMountains",
        datetime(2020, 1, 15),
        None,
        None,
        "NAD27",
        None,
        12,
        7.2,
        66.15,
        "#VALUE!",
        None,
        3,
        None,
        None,
        None,
    ],
]

CM_MODELS = (
    CM_ChemicalData,
    CM_DetectionLimits,
    CM_References,
    CM_MineralSystems,
    CM_WorldComparisons,
    CM_WorldReferences,
    CM_WorkbookMetadata,
)


def _write_workbook(path, *, chem_headers=None, extra_header=None, banner=True):
    """Build a miniature workbook with the same layout as the real one."""
    headers = list(chem_headers if chem_headers is not None else CHEM_HEADERS)
    units = list(CHEM_UNITS)
    if extra_header is not None:
        headers.append(extra_header)
        units.append(None)

    workbook = openpyxl.Workbook()
    chemical_data = workbook.active
    chemical_data.title = "ChemicalData"
    if banner:
        chemical_data.append(["Chemical analyses of samples"])
    chemical_data.append(headers)
    chemical_data.append(["Units"] + units[1:])
    for row in CHEM_ROWS:
        chemical_data.append(row)

    # GIS: same columns, no banner and no units row, and one fewer row --
    # the drift the mirror deliberately preserves.
    gis = workbook.create_sheet("GIS")
    gis.append(headers)
    gis.append(CHEM_ROWS[0])

    # QAQC: header on row 2, no MapSymbol/Pd/Pt, capitalized Latitude/Longitude.
    qaqc = workbook.create_sheet("QAQC")
    dropped = {"MapSymbol", "Pd", "Pt"}
    keep = [i for i, h in enumerate(headers) if h.strip() not in dropped]
    qaqc.append([])
    qaqc.append(
        [
            {"latitude": "Latitude", "longitude": "Longitude"}.get(
                headers[i], headers[i]
            )
            for i in keep
        ]
    )
    qaqc.append([CHEM_ROWS[0][i] if i < len(CHEM_ROWS[0]) else None for i in keep])

    detection_limits = workbook.create_sheet("DetectionLimits")
    detection_limits.append(["Method C_ICPOES_MS-61"])
    detection_limits.append(["Element", "Lower Reporting Limit", "Unit"])
    detection_limits.append(["Ag", 0.01, "ppm"])
    detection_limits.append(["Al", 100, "ppm"])

    references = workbook.create_sheet("References")
    references.append(["McLemore, V.T., 2025, Critical minerals in New Mexico"])
    references.append(["Hofstra, A.H. and Kreiner, D.C., 2020, Mineral systems"])

    mineral_systems = workbook.create_sheet("MineralSystems")
    mineral_systems.append(
        [
            "System Name",
            "Synopsis",
            "Deposit types",
            "Principal commodities",
            "Critical minerals",
            "References",
            "Phase 2",
            "Phase 3",
            "Phase 4",
        ]
    )
    mineral_systems.append(["Magmatic REE", "syn", "Peralkaline", "REE", "Nd", "H&K"])

    world = workbook.create_sheet("world")
    world.append(["area", "deposit", "reference", "La", "Ce", "TREE", "grade %"])
    world.append(["crustal abundance", None, None, 31, 63, 200, 0.1])

    world_ref = workbook.create_sheet("world_ref")
    world_ref.append(["Boyer, D.S., 2011, La Paz Rare Earth Project"])

    general = workbook.create_sheet("General Information")
    general.append(["Critical Minerals in New Mexico"])
    general.append(["Title", "Earth MRI database"])

    metadata = workbook.create_sheet("MetaData")
    metadata.append(["Title", "Earth MRI database"])

    definitions = workbook.create_sheet("DefinitionOfFields")
    definitions.append(["Area", "Geographic area, generally mining district"])

    workbook.save(path)
    return path


def _clear_mirror(session):
    for model in CM_MODELS:
        session.execute(delete(model))
    session.commit()


@pytest.fixture()
def cm_workbook(tmp_path):
    path = _write_workbook(tmp_path / "McLemoreMasterChem_test.xlsx")
    yield path
    with session_ctx() as session:
        _clear_mirror(session)


def _load(path):
    """Load the workbook the way a caller must: the loader does not commit."""
    with session_ctx() as session:
        _clear_mirror(session)
        result = load_cm_workbook(path, session)
        session.commit()
        return result


def test_source_headers_cover_every_mirrored_column():
    """Every CM_ChemicalData data column maps back to a workbook header."""
    provenance = {"id", "source_sheet", "source_row"}
    columns = {c.name for c in CM_ChemicalData.__table__.columns} - provenance
    assert columns == set(SOURCE_HEADER_BY_COLUMN["CM_ChemicalData"])
    assert len(columns) == 118


def test_analyte_units_are_declared_for_every_unit_suffixed_column():
    suffixed = {
        c.name
        for c in CM_ChemicalData.__table__.columns
        if c.name.endswith(("_pct", "_ppm", "_ppb"))
    }
    # tds_mg_l and depth_legnth_ft carry units in their source header, not in
    # the units row, so they are not analytes.
    assert suffixed == set(ANALYTE_UNITS)
    assert ANALYTE_UNITS["au_ppb"] == "ppb"
    assert ANALYTE_UNITS["sio2_pct"] == "%"
    assert ANALYTE_UNITS["as_ppm"] == "ppm"


@pytest.mark.parametrize(
    "value, expected",
    [
        (None, None),
        ("", None),
        ("   ", None),
        ("<0.1", "<0.1"),
        ("#VALUE!", "#VALUE!"),
        ("  ZuniMountains  ", "ZuniMountains"),
        (0, "0"),
        (74.5, "74.5"),
        # A double whose shortest round-trip repr is not its rounded form:
        # the mirror keeps every digit rather than truncating.
        (101.83999999999999, "101.83999999999999"),
        (Decimal("0.010"), "0.010"),
        (True, "TRUE"),
        (datetime(2019, 8, 28), "2019-08-28"),
        (datetime(2019, 8, 28, 13, 45), "2019-08-28 13:45:00"),
        (date(2019, 8, 28), "2019-08-28"),
    ],
)
def test_cell_rendering_preserves_the_source_value(value, expected):
    assert cell_to_text(value) == expected


def test_load_mirrors_every_sheet(cm_workbook):
    result = _load(cm_workbook)

    assert result.rows_by_sheet == {
        "ChemicalData": 2,
        "GIS": 1,
        "QAQC": 1,
        "DetectionLimits": 2,
        "References": 2,
        "MineralSystems": 2,
        "world": 1,
        "world_ref": 1,
        "General Information": 2,
        "MetaData": 1,
        "DefinitionOfFields": 1,
    }
    assert result.total_rows == 16

    with session_ctx() as session:
        assert (
            session.scalar(
                select(CM_DetectionLimits).where(CM_DetectionLimits.element == "Ag")
            ).method
            == "Method C_ICPOES_MS-61"
        )
        assert (
            session.scalars(select(CM_References.citation))
            .all()[0]
            .startswith("McLemore")
        )
        assert (
            session.scalar(
                select(CM_WorldComparisons.grade_pct).where(
                    CM_WorldComparisons.area == "crustal abundance"
                )
            )
            == "0.1"
        )
        assert session.scalar(select(CM_WorldReferences.citation)).startswith("Boyer")
        assert (
            session.scalar(
                select(CM_WorkbookMetadata.value).where(
                    CM_WorkbookMetadata.source_sheet == "DefinitionOfFields",
                    CM_WorkbookMetadata.label == "Area",
                )
            )
            == "Geographic area, generally mining district"
        )
        # The MineralSystems header row is mirrored positionally, notes and all.
        assert (
            session.scalar(
                select(CM_MineralSystems.system_name).where(
                    CM_MineralSystems.source_row == 1
                )
            )
            == "System Name"
        )


def test_all_three_chemistry_sheets_land_in_one_table(cm_workbook):
    _load(cm_workbook)

    with session_ctx() as session:
        sheets = session.scalars(
            select(CM_ChemicalData.source_sheet)
            .distinct()
            .order_by(CM_ChemicalData.source_sheet)
        ).all()
        assert set(sheets) == set(CM_CHEMISTRY_SOURCE_SHEETS)

        # Same sample name, mirrored once per sheet -- no cross-sheet dedupe.
        gal1 = session.scalars(
            select(CM_ChemicalData).where(CM_ChemicalData.sample == "Gal1")
        ).all()
        assert len(gal1) == 3


def test_source_row_records_the_spreadsheet_row_number(cm_workbook):
    _load(cm_workbook)

    with session_ctx() as session:
        # ChemicalData data starts on row 4 (banner, header, units).
        chemical_data = session.scalars(
            select(CM_ChemicalData)
            .where(CM_ChemicalData.source_sheet == "ChemicalData")
            .order_by(CM_ChemicalData.source_row)
        ).all()
        assert [row.source_row for row in chemical_data] == [4, 5]
        # GIS data starts on row 2 (header only).
        assert (
            session.scalar(
                select(CM_ChemicalData.source_row).where(
                    CM_ChemicalData.source_sheet == "GIS"
                )
            )
            == 2
        )
        # QAQC data starts on row 3 (blank row, header).
        assert (
            session.scalar(
                select(CM_ChemicalData.source_row).where(
                    CM_ChemicalData.source_sheet == "QAQC"
                )
            )
            == 3
        )


def test_cells_are_mirrored_without_reinterpretation(cm_workbook):
    _load(cm_workbook)

    with session_ctx() as session:
        row = session.scalar(
            select(CM_ChemicalData).where(
                CM_ChemicalData.source_sheet == "ChemicalData",
                CM_ChemicalData.source_row == 4,
            )
        )
        # Censored values keep their qualifier instead of becoming NULL or 0.
        assert row.au_ppb == "<10"
        assert row.in_ppm == "<0.1"
        assert row.pt_ppm == "<5"
        # Numbers keep their value, not a rounded rendering.
        assert row.total_pct == "101.84"
        assert row.sio2_pct == "74.5"
        # Dates become ISO-8601, not Excel serials.
        assert row.date_collected == "2019-08-28"
        # Blank cells are NULL, not empty strings.
        assert row.paste_ph is None

        excel_error = session.scalar(
            select(CM_ChemicalData.total_pct).where(
                CM_ChemicalData.source_sheet == "ChemicalData",
                CM_ChemicalData.source_row == 5,
            )
        )
        assert excel_error == "#VALUE!"


def test_qaqc_columns_absent_from_the_sheet_are_null(cm_workbook):
    _load(cm_workbook)

    with session_ctx() as session:
        row = session.scalar(
            select(CM_ChemicalData).where(CM_ChemicalData.source_sheet == "QAQC")
        )
        for column in QAQC_MISSING_COLUMNS:
            assert getattr(row, column) is None
        # Capitalized Latitude/Longitude still map to the mirror columns.
        assert row.latitude == "34.19701153"
        assert row.longitude == "-105.74021543"


def test_reload_replaces_rows_per_sheet_without_duplicating(cm_workbook):
    _load(cm_workbook)
    with session_ctx() as session:
        first = load_cm_workbook(cm_workbook, session)
        session.commit()
    with session_ctx() as session:
        second = load_cm_workbook(cm_workbook, session)
        session.commit()

    assert first.rows_by_sheet == second.rows_by_sheet
    with session_ctx() as session:
        assert (
            session.scalar(
                select(CM_ChemicalData.source_row).where(
                    CM_ChemicalData.sample == "Gal1"
                )
            )
            is not None
        )
        assert (
            len(
                session.scalars(
                    select(CM_ChemicalData).where(CM_ChemicalData.sample == "Gal1")
                ).all()
            )
            == 3
        )


def test_row_count_drift_between_chemicaldata_and_gis_is_reported(cm_workbook):
    result = _load(cm_workbook)

    assert any("reconciliation is deferred" in warning for warning in result.warnings)


def test_unknown_header_aborts_the_load(tmp_path):
    path = _write_workbook(
        tmp_path / "extra_column.xlsx", extra_header="NewAnalyteNobodyMapped"
    )

    with pytest.raises(CMWorkbookError, match="NewAnalyteNobodyMapped"):
        _load(path)

    with session_ctx() as session:
        assert session.scalars(select(CM_ChemicalData)).all() == []


def test_moved_header_row_aborts_the_load(tmp_path):
    # Dropping the banner shifts every ChemicalData row up by one.
    path = _write_workbook(tmp_path / "no_banner.xlsx", banner=False)

    with pytest.raises(CMWorkbookError, match="is not the header row"):
        _load(path)


def test_missing_workbook_raises():
    with pytest.raises(CMWorkbookError, match="workbook not found"):
        with session_ctx() as session:
            load_cm_workbook("/nonexistent/McLemoreMasterChem.xlsx", session)
