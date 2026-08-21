# ===============================================================================
# Copyright 2026 ross
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ===============================================================================
"""Build the reconciliation workbook for the critical-minerals mirror.

Reads the delivered McLemore workbook and writes a decision workbook for
V.T. McLemore: where the ChemicalData and GIS sheets disagree, which rows exist
in only one of them, and which values fail a data-integrity check (units that
contradict the declared unit, oxide totals that do not add up, coordinates
outside New Mexico, non-numeric analyte text, inconsistent detection limits,
ambiguous sample names).

Nothing here writes to the database and nothing is decided here -- every
judgement column is left blank for a human. The CM_legacy mirror stores both
sheets verbatim in the meantime; see docs/critical-minerals-legacy-mirror.md.

    python -m scripts.cm_reconciliation_report --workbook <path> --out <path>
"""

from __future__ import annotations

import argparse
import difflib
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable, Sequence

from db.cm_legacy import ANALYTE_UNITS, SOURCE_HEADER_BY_COLUMN
from services.cm_legacy_mirror import (
    CHEMISTRY_SHEET_LAYOUT,
    cell_to_text,
    column_by_header,
    normalize_header,
)

# New Mexico bounding box, generous by a few minutes on every side.
NM_LATITUDE = (31.20, 37.05)
NM_LONGITUDE = (-109.10, -102.95)

# Oxides the workbook's own "Total" column is meant to sum.
MAJOR_OXIDES = (
    "sio2_pct",
    "tio2_pct",
    "al2o3_pct",
    "fe2o3t_pct",
    "mno_pct",
    "mgo_pct",
    "cao_pct",
    "na2o_pct",
    "k2o_pct",
    "p2o5_pct",
    "loi_pct",
)
# Oxide totals this far from the sum of the majors are called out. A major-
# element analysis is normally accepted at 100 +/- a couple of percent.
TOTAL_TOLERANCE = 5.0
MINIMUM_MAJORS_FOR_TOTAL_CHECK = 8
TOTAL_PLAUSIBLE_RANGE = (95.0, 105.0)
# An analyte with more distinct censoring thresholds than this is being
# reported against detection limits that vary by lab or by decade.
THRESHOLD_SPREAD_FLAG = 3

NUMERIC = re.compile(r"^[<>]?\s*-?(\d+\.?\d*|\.\d+)([eE][-+]?\d+)?$")
ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}")

DECISION_CHOICES = (
    '"ChemicalData,GIS,Most recent value,Case by case,Neither - fix source"'
)
INTEGRITY_CHOICES = (
    '"Fix in source workbook,Handle in transform,Accept as-is,Needs review"'
)


@dataclass(frozen=True)
class Value:
    """A parsed workbook cell."""

    text: str | None
    number: float | None
    kind: str | None  # "numeric" | "censored" | "text" | None


def parse_value(cell: Any) -> Value:
    text = cell_to_text(cell)
    if text is None:
        return Value(None, None, None)
    candidate = text.replace(",", "").strip()
    if NUMERIC.match(candidate):
        number = float(candidate.lstrip("<> "))
        return Value(text, number, "censored" if candidate[0] in "<>" else "numeric")
    return Value(text, None, "text")


@dataclass
class Sheet:
    """A chemistry sheet flattened to {column: text} records."""

    name: str
    records: list[dict[str, str | None]]
    row_numbers: list[int]


def read_sheet(workbook: Any, name: str, columns: dict[str, str]) -> Sheet:
    if name not in workbook.sheetnames:
        raise ValueError(f"workbook is missing the {name!r} sheet")

    header_row_number, first_data_row_number = CHEMISTRY_SHEET_LAYOUT[name]
    rows = list(workbook[name].iter_rows(values_only=True))
    if len(rows) < header_row_number:
        raise ValueError(f"sheet {name!r} has no header row {header_row_number}")

    header = rows[header_row_number - 1]
    index_to_column: dict[int, str] = {}
    unknown: list[str] = []
    for index, cell in enumerate(header):
        text = cell_to_text(cell)
        if text is None:
            continue
        column = columns.get(normalize_header(text))
        if column is None:
            unknown.append(text)
        else:
            index_to_column[index] = column

    if unknown:
        raise ValueError(f"sheet {name!r} has unmapped header(s): {unknown[:10]}")

    records: list[dict[str, str | None]] = []
    row_numbers: list[int] = []
    for offset, row in enumerate(rows[first_data_row_number - 1 :]):
        if all(cell_to_text(cell) is None for cell in row):
            continue
        records.append(
            {
                column: cell_to_text(row[index]) if index < len(row) else None
                for index, column in index_to_column.items()
            }
        )
        row_numbers.append(first_data_row_number + offset)
    return Sheet(name, records, row_numbers)


def align(
    left: Sheet, right: Sheet
) -> tuple[list[tuple[int, int]], list[int], list[int]]:
    """Line the two sheets up on their sample-name sequence.

    Returns aligned (left index, right index) pairs plus the indexes that exist
    on only one side.
    """
    left_keys = [(record.get("sample") or "").strip() for record in left.records]
    right_keys = [(record.get("sample") or "").strip() for record in right.records]
    matcher = difflib.SequenceMatcher(a=left_keys, b=right_keys, autojunk=False)

    pairs: list[tuple[int, int]] = []
    left_only: list[int] = []
    right_only: list[int] = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            pairs.extend(zip(range(i1, i2), range(j1, j2)))
        else:
            left_only.extend(range(i1, i2))
            right_only.extend(range(j1, j2))
    return pairs, left_only, right_only


def classify(left_text: str | None, right_text: str | None) -> str | None:
    """Name the kind of disagreement between two cells, or None if they agree."""
    left = (left_text or "").strip()
    right = (right_text or "").strip()
    if left == right:
        return None
    if not right:
        return "only in ChemicalData"
    if not left:
        return "only in GIS"
    if right.startswith("#"):
        return "Excel error in GIS"
    if left.startswith("#"):
        return "Excel error in ChemicalData"
    left_value, right_value = parse_value(left), parse_value(right)
    if left_value.number is not None and right_value.number is not None:
        if left_value.number == right_value.number:
            return "same number, different text"
        return "different number"
    return "different text"


def compare_sheets(
    chemical_data: Sheet, gis: Sheet, columns: Sequence[str]
) -> tuple[list[dict[str, Any]], Counter, Counter]:
    pairs, chemical_data_only, _ = align(chemical_data, gis)
    differences: list[dict[str, Any]] = []
    by_column: Counter = Counter()
    by_column_kind: Counter = Counter()

    for left_index, right_index in pairs:
        left_record = chemical_data.records[left_index]
        right_record = gis.records[right_index]
        for column in columns:
            kind = classify(left_record.get(column), right_record.get(column))
            if kind is None:
                continue
            by_column[column] += 1
            by_column_kind[(column, kind)] += 1
            differences.append(
                {
                    "sample": left_record.get("sample"),
                    "area": left_record.get("area"),
                    "column": column,
                    "header": SOURCE_HEADER_BY_COLUMN["CM_ChemicalData"][column],
                    "chemical_data_value": left_record.get(column),
                    "gis_value": right_record.get(column),
                    "kind": kind,
                    "chemical_data_row": chemical_data.row_numbers[left_index],
                    "gis_row": gis.row_numbers[right_index],
                }
            )
    return differences, by_column, by_column_kind


def recommend(column: str, kinds: dict[str, int]) -> str:
    """Suggest a starting point for each column, to be confirmed or overruled."""
    only_gis = kinds.get("only in GIS", 0)
    only_chemical_data = kinds.get("only in ChemicalData", 0)
    excel_errors = kinds.get("Excel error in GIS", 0) + kinds.get(
        "Excel error in ChemicalData", 0
    )
    conflicts = (
        kinds.get("different number", 0)
        + kinds.get("different text", 0)
        + kinds.get("same number, different text", 0)
    )
    if conflicts:
        return "Conflicting values - needs a per-row ruling"
    if excel_errors and not (only_gis or only_chemical_data):
        return "Take the sheet without the #VALUE! error"
    if only_gis and only_chemical_data:
        return "Fill blanks from whichever sheet has a value"
    if only_gis:
        return "Take GIS (it fills blanks ChemicalData never got)"
    if only_chemical_data:
        return "Take ChemicalData (GIS is missing these)"
    return "No action"


def integrity_findings(
    sheet: Sheet, detection_limits: dict[str, float]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Run the data-integrity checks over one sheet.

    Returns (per-cell/row detail, threshold spread rows, duplicate-name rows).
    """
    detail: list[dict[str, Any]] = []

    def flag(record, row_number, issue, column, value, note):
        detail.append(
            {
                "sheet": sheet.name,
                "source_row": row_number,
                "sample": record.get("sample"),
                "area": record.get("area"),
                "reference": record.get("reference"),
                "issue": issue,
                "column": column,
                "value": value,
                "note": note,
            }
        )

    thresholds: dict[str, set[float]] = defaultdict(set)

    for record, row_number in zip(sheet.records, sheet.row_numbers):
        # --- analyte values against the unit the workbook declares for them
        for column, unit in ANALYTE_UNITS.items():
            value = parse_value(record.get(column))
            if value.kind is None:
                continue
            if value.kind == "censored":
                thresholds[column].add(value.number)
            if value.kind == "text":
                flag(
                    record,
                    row_number,
                    "Non-numeric analyte value",
                    column,
                    value.text,
                    "Not a number and not a '<' detection limit; the transform "
                    "cannot store it as a result",
                )
                continue
            if (
                unit == "%"
                and column != "total_pct"
                and value.number is not None
                and value.number > 100
            ):
                flag(
                    record,
                    row_number,
                    "Value impossible for the declared unit",
                    column,
                    value.text,
                    f"Declared as % but reported {value.number:g}; looks like ppm",
                )
            if (
                unit == "ppb"
                and value.kind == "censored"
                and value.number is not None
                and value.number < 0.01
            ):
                flag(
                    record,
                    row_number,
                    "Detection limit implausible for the declared unit",
                    column,
                    value.text,
                    f"Declared as ppb but censored at {value.number:g}; looks like ppm",
                )

        # --- oxide totals
        total = parse_value(record.get("total_pct"))
        majors = [parse_value(record.get(column)) for column in MAJOR_OXIDES]
        reported = [value.number for value in majors if value.kind == "numeric"]
        if total.kind == "numeric":
            summed = sum(reported)
            enough_majors = len(reported) >= MINIMUM_MAJORS_FOR_TOTAL_CHECK
            if enough_majors and abs(summed - total.number) > TOTAL_TOLERANCE:
                flag(
                    record,
                    row_number,
                    "Total disagrees with the sum of the major oxides",
                    "total_pct",
                    total.text,
                    f"Majors sum to {summed:.2f} from {len(reported)} oxides, "
                    f"a difference of {summed - total.number:+.2f}",
                )
            elif not (
                TOTAL_PLAUSIBLE_RANGE[0] <= total.number <= TOTAL_PLAUSIBLE_RANGE[1]
            ):
                flag(
                    record,
                    row_number,
                    "Total outside 95-105%",
                    "total_pct",
                    total.text,
                    (
                        "Consistent with the majors, so the analysis itself is "
                        "incomplete or the sample is not a whole rock"
                        if enough_majors
                        else f"Only {len(reported)} major oxides reported, so the "
                        "Total cannot be checked against them"
                    ),
                )
        elif total.kind == "text":
            flag(
                record,
                row_number,
                "Total is not a number",
                "total_pct",
                total.text,
                "Broken spreadsheet formula",
            )

        # --- coordinates
        latitude = parse_value(record.get("latitude"))
        longitude = parse_value(record.get("longitude"))
        datum = (record.get("coordinate_system") or "").strip()
        has_latitude = latitude.number is not None
        has_longitude = longitude.number is not None
        if not has_latitude and not has_longitude:
            flag(
                record,
                row_number,
                "No coordinates",
                "latitude/longitude",
                None,
                f"Coordinate system is {datum!r}" if datum else "No datum either",
            )
        elif has_latitude != has_longitude:
            flag(
                record,
                row_number,
                "Only one of latitude/longitude",
                "latitude/longitude",
                f"{latitude.text} / {longitude.text}",
                "Unusable as a point",
            )
        else:
            if not datum:
                flag(
                    record,
                    row_number,
                    "Coordinates with no declared datum",
                    "coordinate_system",
                    f"{latitude.text} / {longitude.text}",
                    "Cannot be reprojected; NAD27 and WGS84 differ by ~100 m here",
                )
            if longitude.number > 0:
                flag(
                    record,
                    row_number,
                    "Positive longitude",
                    "longitude",
                    longitude.text,
                    "Missing the minus sign puts the sample in Asia",
                )
            if not NM_LATITUDE[0] <= latitude.number <= NM_LATITUDE[1]:
                flag(
                    record,
                    row_number,
                    "Latitude outside New Mexico",
                    "latitude",
                    latitude.text,
                    f"Outside {NM_LATITUDE[0]}-{NM_LATITUDE[1]}",
                )
            if not NM_LONGITUDE[0] <= longitude.number <= NM_LONGITUDE[1]:
                flag(
                    record,
                    row_number,
                    "Longitude outside New Mexico",
                    "longitude",
                    longitude.text,
                    f"Outside {NM_LONGITUDE[0]} to {NM_LONGITUDE[1]}",
                )

        # --- dates
        collected = (record.get("date_collected") or "").strip()
        analyzed = (record.get("date_analyzed") or "").strip()
        for label, value in (
            ("date_collected", collected),
            ("date_analyzed", analyzed),
        ):
            if value and not ISO_DATE.match(value):
                flag(
                    record,
                    row_number,
                    "Date is not a full date",
                    label,
                    value,
                    "Year only or free text; stored as text, not a date",
                )
        if (
            collected
            and analyzed
            and ISO_DATE.match(collected)
            and ISO_DATE.match(analyzed)
            and analyzed[:10] < collected[:10]
        ):
            flag(
                record,
                row_number,
                "Analyzed before collected",
                "date_analyzed",
                f"collected {collected[:10]}, analyzed {analyzed[:10]}",
                "One of the two dates is wrong",
            )

    threshold_rows = [
        {
            "column": column,
            "header": SOURCE_HEADER_BY_COLUMN["CM_ChemicalData"][column],
            "declared_unit": ANALYTE_UNITS[column],
            "distinct_thresholds": len(values),
            "lowest": min(values),
            "highest": max(values),
            "detection_limits_sheet": detection_limits.get(
                SOURCE_HEADER_BY_COLUMN["CM_ChemicalData"][column].strip()
            ),
            "thresholds": ", ".join(f"{value:g}" for value in sorted(values)[:25]),
        }
        for column, values in sorted(thresholds.items())
        if len(values) > THRESHOLD_SPREAD_FLAG
    ]
    threshold_rows.sort(key=lambda row: -row["distinct_thresholds"])

    names = Counter((record.get("sample") or "").strip() for record in sheet.records)
    duplicate_rows = [
        {
            "sample": name,
            "rows": count,
            "areas": ", ".join(
                sorted(
                    {
                        (record.get("area") or "?").strip()
                        for record in sheet.records
                        if (record.get("sample") or "").strip() == name
                    }
                )
            ),
            "source_rows": ", ".join(
                str(number)
                for record, number in zip(sheet.records, sheet.row_numbers)
                if (record.get("sample") or "").strip() == name
            ),
        }
        for name, count in names.most_common()
        if name and count > 1
    ]

    return detail, threshold_rows, duplicate_rows


def read_detection_limits(workbook: Any) -> dict[str, float]:
    limits: dict[str, float] = {}
    for row in list(workbook["DetectionLimits"].iter_rows(values_only=True))[2:]:
        element = cell_to_text(row[0]) if row else None
        value = parse_value(row[1]) if len(row) > 1 else Value(None, None, None)
        if element and value.number is not None:
            limits[element.strip()] = value.number
    return limits


# ----------------------------------------------------------------------
# workbook writing
# ----------------------------------------------------------------------

HEADER_FILL = "FF1F3B4D"
NOTE_FILL = "FFF2F2F2"
DECISION_FILL = "FFFFF3C4"


def _style_header(
    worksheet: Any, columns: Sequence[str], widths: Sequence[int]
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    worksheet.append(list(columns))
    for index, (column, width) in enumerate(zip(columns, widths), start=1):
        cell = worksheet.cell(row=1, column=index)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.column_dimensions[
            worksheet.cell(row=1, column=index).column_letter
        ].width = width
    worksheet.freeze_panes = "A2"


def _add_table(
    workbook: Any,
    title: str,
    columns: Sequence[str],
    widths: Sequence[int],
    rows: Iterable[Sequence[Any]],
    decision_columns: Sequence[tuple[int, str]] = (),
) -> Any:
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    worksheet = workbook.create_sheet(title)
    _style_header(worksheet, columns, widths)
    count = 0
    for row in rows:
        worksheet.append(list(row))
        count += 1
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=len(columns)).column_letter}{max(count + 1, 2)}"

    for column_index, choices in decision_columns:
        letter = worksheet.cell(row=1, column=column_index).column_letter
        validation = DataValidation(type="list", formula1=choices, allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{max(count + 1, 2)}")
        for row_index in range(2, count + 2):
            worksheet.cell(row=row_index, column=column_index).fill = PatternFill(
                "solid", fgColor=DECISION_FILL
            )
        worksheet.cell(row=1, column=column_index).alignment = Alignment(
            vertical="center", wrap_text=True
        )
    return worksheet


def _add_readme(workbook: Any, facts: Sequence[tuple[str, str]]) -> None:
    from openpyxl.styles import Alignment, Font

    worksheet = workbook.create_sheet("README", 0)
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 104
    for label, value in facts:
        worksheet.append([label, value])
        row = worksheet.max_row
        worksheet.cell(row=row, column=1).font = Font(bold=not label.startswith(" "))
        worksheet.cell(row=row, column=1).alignment = Alignment(vertical="top")
        worksheet.cell(row=row, column=2).alignment = Alignment(
            vertical="top", wrap_text=True
        )


def build_report(workbook_path: Path, output_path: Path) -> dict[str, int]:
    import openpyxl

    source = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    columns = column_by_header("CM_ChemicalData")
    all_columns = list(SOURCE_HEADER_BY_COLUMN["CM_ChemicalData"])

    chemical_data = read_sheet(source, "ChemicalData", columns)
    gis = read_sheet(source, "GIS", columns)
    detection_limits = read_detection_limits(source)

    differences, by_column, by_column_kind = compare_sheets(
        chemical_data, gis, all_columns
    )
    pairs, chemical_data_only, gis_only = align(chemical_data, gis)
    detail, thresholds, duplicates = integrity_findings(chemical_data, detection_limits)
    source.close()

    kinds_by_column: dict[str, dict[str, int]] = defaultdict(dict)
    for (column, kind), count in by_column_kind.items():
        kinds_by_column[column][kind] = count

    report = openpyxl.Workbook()
    report.remove(report.active)

    rows_with_differences = len(
        {(row["chemical_data_row"], row["gis_row"]) for row in differences}
    )
    issue_counts = Counter(row["issue"] for row in detail)

    _add_table(
        report,
        "ColumnDecisions",
        [
            "Mirror column",
            "Workbook header",
            "Cells differing",
            "Only in GIS",
            "Only in ChemicalData",
            "Different value",
            "Excel error",
            "Suggested starting point",
            "DECISION",
            "Notes",
        ],
        [24, 34, 12, 12, 14, 14, 12, 46, 22, 46],
        (
            [
                column,
                SOURCE_HEADER_BY_COLUMN["CM_ChemicalData"][column],
                by_column[column],
                kinds_by_column[column].get("only in GIS", 0),
                kinds_by_column[column].get("only in ChemicalData", 0),
                kinds_by_column[column].get("different number", 0)
                + kinds_by_column[column].get("different text", 0)
                + kinds_by_column[column].get("same number, different text", 0),
                kinds_by_column[column].get("Excel error in GIS", 0)
                + kinds_by_column[column].get("Excel error in ChemicalData", 0),
                recommend(column, kinds_by_column[column]),
                None,
                None,
            ]
            for column, _ in by_column.most_common()
        ),
        decision_columns=[(9, DECISION_CHOICES)],
    )

    _add_table(
        report,
        "CellDifferences",
        [
            "SAMPLE",
            "Area",
            "Mirror column",
            "Workbook header",
            "ChemicalData value",
            "GIS value",
            "Kind of difference",
            "ChemicalData row",
            "GIS row",
            "DECISION",
        ],
        [20, 18, 22, 32, 24, 24, 26, 16, 12, 22],
        (
            [
                row["sample"],
                row["area"],
                row["column"],
                row["header"],
                row["chemical_data_value"],
                row["gis_value"],
                row["kind"],
                row["chemical_data_row"],
                row["gis_row"],
                None,
            ]
            for row in sorted(
                differences, key=lambda row: (row["column"], row["sample"] or "")
            )
        ),
        decision_columns=[(10, DECISION_CHOICES)],
    )

    _add_table(
        report,
        "RowsOnlyInOneSheet",
        [
            "Sheet",
            "SAMPLE",
            "Area",
            "Reference",
            "Source row",
            "Has coordinates",
            "Note",
            "DECISION",
        ],
        [16, 22, 20, 40, 12, 16, 46, 22],
        (
            [
                sheet.name,
                sheet.records[index].get("sample"),
                sheet.records[index].get("area"),
                sheet.records[index].get("reference"),
                sheet.row_numbers[index],
                (
                    "yes"
                    if parse_value(sheet.records[index].get("latitude")).number
                    is not None
                    else "no"
                ),
                note,
                None,
            ]
            for sheet, indexes, note in (
                (
                    chemical_data,
                    chemical_data_only,
                    "In ChemicalData only - appended after GIS was last synced",
                ),
                (gis, gis_only, "In GIS only"),
            )
            for index in indexes
        ),
        decision_columns=[
            (8, '"Add to GIS,Drop - not a sample,Keep as-is,Needs review"')
        ],
    )

    _add_table(
        report,
        "IntegritySummary",
        [
            "Issue",
            "Rows affected",
            "Why it matters",
            "RESOLUTION",
            "Notes",
        ],
        [46, 14, 76, 26, 40],
        (
            [issue, count, INTEGRITY_NOTES.get(issue, ""), None, None]
            for issue, count in issue_counts.most_common()
        ),
        decision_columns=[(4, INTEGRITY_CHOICES)],
    )

    _add_table(
        report,
        "IntegrityDetail",
        [
            "Sheet",
            "Source row",
            "SAMPLE",
            "Area",
            "Reference",
            "Issue",
            "Column",
            "Value",
            "Detail",
            "RESOLUTION",
        ],
        [14, 12, 20, 18, 34, 40, 20, 20, 60, 24],
        (
            [
                row["sheet"],
                row["source_row"],
                row["sample"],
                row["area"],
                row["reference"],
                row["issue"],
                row["column"],
                row["value"],
                row["note"],
                None,
            ]
            for row in sorted(detail, key=lambda row: (row["issue"], row["source_row"]))
        ),
        decision_columns=[(10, INTEGRITY_CHOICES)],
    )

    _add_table(
        report,
        "DetectionLimitSpread",
        [
            "Mirror column",
            "Workbook header",
            "Declared unit",
            "Distinct '<' thresholds",
            "Lowest",
            "Highest",
            "DetectionLimits sheet",
            "Thresholds seen (first 25)",
            "RESOLUTION",
        ],
        [22, 24, 14, 20, 12, 12, 20, 62, 24],
        (
            [
                row["column"],
                row["header"],
                row["declared_unit"],
                row["distinct_thresholds"],
                row["lowest"],
                row["highest"],
                row["detection_limits_sheet"],
                row["thresholds"],
                None,
            ]
            for row in thresholds
        ),
        decision_columns=[(9, INTEGRITY_CHOICES)],
    )

    _add_table(
        report,
        "DuplicateSampleNames",
        ["SAMPLE", "Rows with this name", "Areas", "Source rows", "RESOLUTION"],
        [22, 20, 46, 40, 26],
        (
            [row["sample"], row["rows"], row["areas"], row["source_rows"], None]
            for row in duplicates
        ),
        decision_columns=[
            (5, '"Same sample - de-duplicate,Different samples - rename,Needs review"')
        ],
    )

    _add_readme(
        report,
        [
            ("Critical minerals reconciliation", ""),
            (
                "What this is",
                "Everywhere the delivered workbook contradicts itself, laid out for a "
                "decision. Two kinds of problem: the ChemicalData and GIS sheets "
                "disagree with each other, and some values fail a data-integrity "
                "check regardless of which sheet they came from.",
            ),
            (
                "Who fills it in",
                "V.T. McLemore, or whoever owns the source data. Every yellow column "
                "is blank on purpose and has a dropdown; the grey columns are "
                "generated and should not be edited.",
            ),
            (
                "What happens next",
                "The CM_legacy staging mirror already holds BOTH sheets verbatim, so "
                "nothing has been lost and nothing has been merged. The rulings here "
                "become the merge rules for the transform into the Ocotillo data "
                "model.",
            ),
            ("", ""),
            ("Source workbook", workbook_path.name),
            ("Generated", date.today().isoformat()),
            ("ChemicalData rows", str(len(chemical_data.records))),
            ("GIS rows", str(len(gis.records))),
            ("Rows aligned on SAMPLE", str(len(pairs))),
            ("Rows that disagree", str(rows_with_differences)),
            ("Cells that disagree", str(len(differences))),
            ("Integrity findings", str(len(detail))),
            ("", ""),
            ("Sheet: ColumnDecisions", "One row per column that differs. Start here."),
            (
                "Sheet: CellDifferences",
                "Every disagreeing cell, so a column-level ruling can be checked or "
                "overridden row by row.",
            ),
            (
                "Sheet: RowsOnlyInOneSheet",
                "Samples present in one sheet and not the other. ChemicalData has "
                "rows appended after GIS was last synced; one of them is not a "
                "sample at all but the 'NOTE: SEE THE ORIGINAL CITATION...' text.",
            ),
            (
                "Sheet: IntegritySummary",
                "Data-integrity findings grouped by issue, worst first.",
            ),
            ("Sheet: IntegrityDetail", "The individual rows behind each finding."),
            (
                "Sheet: DetectionLimitSpread",
                "Analytes reported against many different '<' detection limits, i.e. "
                "results pooled across labs and decades without normalizing.",
            ),
            (
                "Sheet: DuplicateSampleNames",
                "SAMPLE is not unique, so it cannot be the key on its own.",
            ),
            ("", ""),
            (
                "Note on units",
                "Integrity checks compare each analyte against the unit declared in "
                "the workbook's own units row (ChemicalData row 3). A '%' column "
                "holding 27700 is the clearest case: that is ppm in a percent "
                "column.",
            ),
            (
                "Note on scope",
                "Only ChemicalData is integrity-checked. Checking GIS as well would "
                "double every finding without adding information, since the "
                "sheet-to-sheet differences are already listed above.",
            ),
        ],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    report.save(output_path)

    return {
        "chemical_data_rows": len(chemical_data.records),
        "gis_rows": len(gis.records),
        "aligned_rows": len(pairs),
        "rows_with_differences": rows_with_differences,
        "cell_differences": len(differences),
        "columns_with_differences": len(by_column),
        "rows_only_in_chemical_data": len(chemical_data_only),
        "rows_only_in_gis": len(gis_only),
        "integrity_findings": len(detail),
        "integrity_issue_types": len(issue_counts),
        "analytes_with_threshold_spread": len(thresholds),
        "duplicate_sample_names": len(duplicates),
    }


INTEGRITY_NOTES = {
    "Value impossible for the declared unit": (
        "The workbook declares this column's unit in its units row, and the value "
        "cannot be that unit -- a percentage over 100. Almost certainly ppm "
        "reported in a % column. The transform cannot guess which rows to convert."
    ),
    "Detection limit implausible for the declared unit": (
        "Column is declared ppb but the '<' limit is a ppm-scale number, so some "
        "rows in this column are in different units from the rest."
    ),
    "Non-numeric analyte value": (
        "Text such as 'bd', 'nd', 'tr', 'nr', 'n/a', '----' or '>2%' where a number "
        "belongs. Each token needs a ruling: below detection (and at what limit), "
        "not determined, trace, or not reported."
    ),
    "Total disagrees with the sum of the major oxides": (
        "The reported Total is not the sum of SiO2..LOI for that row, so either a "
        "major oxide is missing from the row or the Total is stale."
    ),
    "Total outside 95-105%": (
        "Total is far from 100%, so the analysis is partial rather than a complete "
        "whole-rock analysis. Fine for trace-element work, misleading if read as "
        "whole rock. Excluded from the unit check above because Total is a sum, "
        "not a measurement."
    ),
    "Total is not a number": (
        "A broken spreadsheet formula ('#VALUE!') rather than a measurement."
    ),
    "No coordinates": (
        "The sample cannot be mapped or matched to a Location. Some of these rows "
        "still declare a coordinate system, which suggests the coordinates were "
        "meant to be filled in."
    ),
    "Only one of latitude/longitude": "Half a coordinate pair is unusable as a point.",
    "Coordinates with no declared datum": (
        "NAD27 and WGS84 differ by roughly 100 m in New Mexico, so an undeclared "
        "datum is a 100 m position error."
    ),
    "Positive longitude": "A missing minus sign puts the sample on the other side of the world.",
    "Latitude outside New Mexico": (
        "Outside the state bounding box: transposed digits, a swapped lat/long pair, "
        "or a genuinely out-of-state sample that should be labelled as such."
    ),
    "Longitude outside New Mexico": (
        "Outside the state bounding box -- same causes as an out-of-range latitude."
    ),
    "Date is not a full date": (
        "Year-only or free-text dates cannot become a field-event date without a "
        "convention for what day to use."
    ),
    "Analyzed before collected": "One of the two dates is wrong.",
}


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        required=True,
        type=Path,
        help="Path to the McLemoreMasterChem .xlsx",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("cm_reconciliation.xlsx"),
        help="Where to write the reconciliation workbook",
    )
    arguments = parser.parse_args(argv)

    summary = build_report(arguments.workbook, arguments.out)
    width = max(len(key) for key in summary)
    for key, value in summary.items():
        print(f"{key:<{width}} {value:>8}")
    print(f"\nwrote {arguments.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
