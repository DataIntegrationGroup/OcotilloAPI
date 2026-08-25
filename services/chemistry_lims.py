# ===============================================================================
# Copyright 2026 ross
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ===============================================================================
"""
Ingest a laboratory LIMS chemistry workbook into the legacy NMA chemistry
tables (NMA_MajorChemistry / NMA_MinorTraceChemistry).

Ported from the AMPAPI ``chemfile.py`` ingestion script. The AMPAPI original
read an ``.xls`` LIMS export with ``xlrd`` and inserted rows into the SQL Server
``MajorChemistry`` / ``MinorandTraceChemistry`` tables keyed off the
``Chemistry SampleInfo`` table. This adaptation:

* reads an ``.xlsx`` workbook with ``openpyxl``,
* maps each LIMS ``Param`` to an analyte code + target table via
  :func:`lookup_analyte`,
* resolves each ``SamplePointID`` (the base well PointID) to a ``Thing`` by
  name,
* appends each distinct lab sample (``WCLab_ID``) as a new
  ``NMA_Chemistry_SampleInfo`` row whose ``nma_sample_point_id`` is the base
  PointID with the next letter incrementor appended (``A``, ``B``, ... ``Z``,
  ``AA``, ...), skipping a lab sample already recorded for that well.

The public entrypoint is :func:`bulk_upload_chemistry`.
"""

from __future__ import annotations

import io
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from itertools import groupby
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from db import (
    NMA_Chemistry_SampleInfo,
    NMA_MajorChemistry,
    NMA_MinorTraceChemistry,
    Thing,
)
from db.engine import session_ctx

# --- analyte mapping (ported verbatim from AMPAPI chemfile.py) -----------------

MINOR = "MinorandTraceChemistry"
MAJOR = "MajorChemistry"
EPM = "epm"
MGL = "mg/L"
PDIFF = "%Diff"
PH = "pH"
COND = "µS/cm"

ANALYSES_AGENCY = "NMBGMR"


@dataclass(frozen=True)
class AnalyteMapping:
    """Maps a LIMS ``Param`` name to its analyte code and target table.

    ``units`` overrides the LIMS-reported units when set; ``method`` is appended
    to the LIMS analysis method when set.
    """

    lims_param: str
    analyte: str
    table: str
    units: str | None = None
    method: str | None = None


# Every known LIMS ``Param`` -> analyte mapping (ported from AMPAPI chemfile.py).
_ANALYTE_MAPPINGS: list[AnalyteMapping] = [
    AnalyteMapping("alkalinity as caco3", "ALK", MAJOR, method="As CaCO3"),
    AnalyteMapping("aluminum", "Al", MINOR),
    AnalyteMapping("anions total", "TAn", MAJOR, EPM),
    AnalyteMapping("antimony 121", "Sb", MINOR),
    AnalyteMapping("antimony 123", "Sb", MINOR),
    AnalyteMapping("antimony", "Sb", MINOR),
    AnalyteMapping("arsenic", "As", MINOR),
    AnalyteMapping("barium", "Ba", MINOR),
    AnalyteMapping("beryllium", "Be", MINOR),
    AnalyteMapping("bicarbonate (hco3)", "HCO3", MAJOR, method="Alkalinity as HC03"),
    AnalyteMapping("boron 11", "B", MINOR),
    AnalyteMapping("boron", "B", MINOR),
    AnalyteMapping("bromide", "Br", MINOR),
    AnalyteMapping("cadmium 111", "Cd", MINOR),
    AnalyteMapping("cadmium", "Cd", MINOR),
    AnalyteMapping("calcium", "Ca", MAJOR),
    AnalyteMapping("carbonate (co3)", "CO3", MAJOR),
    AnalyteMapping("cations total", "TCat", MAJOR, EPM),
    AnalyteMapping("chloride", "Cl", MAJOR),
    AnalyteMapping("chromium", "Cr", MINOR),
    AnalyteMapping("cobalt", "Co", MINOR),
    AnalyteMapping("copper 65", "Cu", MINOR),
    AnalyteMapping("copper", "Cu", MINOR),
    AnalyteMapping("fluoride", "F", MINOR),
    AnalyteMapping("hardness", "HRD", MAJOR, MGL, method="As CaCO3"),
    AnalyteMapping("iron", "Fe", MINOR),
    AnalyteMapping("lead", "Pb", MINOR),
    AnalyteMapping("lithium", "Li", MINOR),
    AnalyteMapping("magnesium", "Mg", MAJOR),
    AnalyteMapping("manganese", "Mn", MINOR),
    AnalyteMapping("mercury", "Hg", MINOR),
    AnalyteMapping("molybdenum 95", "Mo", MINOR),
    AnalyteMapping("molybdenum", "Mo", MINOR),
    AnalyteMapping("nickel", "Ni", MINOR),
    AnalyteMapping("nitrate", "NO3", MINOR),
    AnalyteMapping("nitrite", "NO2", MINOR),
    AnalyteMapping("phosphate", "PO4", MINOR),
    AnalyteMapping("percent difference", "IONBAL", MAJOR, PDIFF),
    AnalyteMapping("potassium", "K", MAJOR),
    AnalyteMapping("selenium", "Se", MINOR),
    AnalyteMapping("siliconDioxide", "SiO2", MINOR),
    AnalyteMapping("sio2", "SiO2", MINOR),
    AnalyteMapping("silicon", "Si", MINOR),
    AnalyteMapping("silver 107", "Ag", MINOR),
    AnalyteMapping("silver", "Ag", MINOR),
    AnalyteMapping("sodium", "Na", MAJOR),
    AnalyteMapping("specific conductance", "CONDLAB", MAJOR, COND),
    AnalyteMapping("strontium", "Sr", MINOR),
    AnalyteMapping("sulfate", "SO4", MAJOR),
    AnalyteMapping("tds calc", "TDS", MAJOR, method="Calculation"),
    AnalyteMapping("thallium", "Tl", MINOR),
    AnalyteMapping("thorium", "Th", MINOR),
    AnalyteMapping("tin", "Sn", MINOR),
    AnalyteMapping("titanium", "Ti", MINOR),
    AnalyteMapping("uranium", "U", MINOR),
    AnalyteMapping("vanadium", "V", MINOR),
    AnalyteMapping("zinc 66", "Zn", MINOR),
    AnalyteMapping("zinc", "Zn", MINOR),
    AnalyteMapping("pH", "pHL", MAJOR, PH),
    AnalyteMapping("ortho phosphate", "PO4", MINOR),
]

# Case-insensitive lookup by LIMS ``Param`` name.
_ANALYTE_BY_PARAM: dict[str, AnalyteMapping] = {
    m.lims_param.lower(): m for m in _ANALYTE_MAPPINGS
}


def lookup_analyte(param: str | None) -> AnalyteMapping | None:
    """Return the mapping for a LIMS ``Param`` name, or ``None`` if unknown."""
    if param is None:
        return None
    return _ANALYTE_BY_PARAM.get(str(param).strip().lower())


# Target ORM model per analyte table bucket.
_TABLE_MODEL = {MAJOR: NMA_MajorChemistry, MINOR: NMA_MinorTraceChemistry}


class ChemistryMappingError(Exception):
    """A LIMS row could not be normalized into an analyte measurement."""


@dataclass
class ChemistryUploadResult:
    exit_code: int
    stderr: str
    payload: dict[str, Any]


# --- workbook parsing ----------------------------------------------------------

# Columns the LIMS export is expected to carry. Extra columns are ignored;
# missing columns simply read back as ``None``.
LIMS_COLUMNS = (
    "Param",
    "Results_Units",
    "Dilution",
    "AnalysisTime",
    "SampleNumber",
    "CustomerSampleNumber",
    "SamplePointID",
    "Method",
    "Test",
    "ReportedND",
    "LowerLimit",
    "SampleDate",
)


def read_lims_xlsx(
    source: Path | str | bytes | BinaryIO, sheet_index: int = 0
) -> list[dict]:
    """Read a LIMS ``.xlsx`` workbook into a list of header->value dicts."""
    if isinstance(source, (bytes, bytearray)):
        handle: Any = io.BytesIO(source)
    elif isinstance(source, (str, Path)):
        handle = source
    else:
        handle = source

    wb = load_workbook(filename=handle, read_only=True, data_only=True)
    try:
        sheet = wb.worksheets[sheet_index]
        rows = sheet.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(rows)]
        except StopIteration:
            return []
        records = []
        for row in rows:
            if all(v is None for v in row):
                continue
            records.append(dict(zip(header, row)))
        return records
    finally:
        wb.close()


# --- record normalization ------------------------------------------------------


def _get(record: dict, key: str) -> Any:
    value = record.get(key)
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    return value


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return None


def prep_record(record: dict) -> dict:
    """Normalize one raw LIMS row into an analyte-measurement dict.

    Raises :class:`ChemistryMappingError` when the row cannot be mapped.
    """
    param = _get(record, "Param")
    mapping = lookup_analyte(param)
    if mapping is None:
        raise ChemistryMappingError(f"Unmapped analyte Param={param!r}")

    pointid = _get(record, "SamplePointID") or _get(record, "CustomerSampleNumber")
    if not pointid:
        raise ChemistryMappingError("Missing SamplePointID")

    # The WCLab_ID is the only thing that makes a re-ingest recognizable, so a
    # row without one is not loadable data: it would be appended again under a
    # fresh lettered sample point on every run. Same policy as a missing
    # SamplePointID, a row error that aborts the file before anything is written.
    wclab_id = _get(record, "SampleNumber")
    if not wclab_id:
        raise ChemistryMappingError("Missing SampleNumber")

    units = mapping.units or _get(record, "Results_Units")

    reported = _get(record, "ReportedND")
    if reported is not None and str(reported).upper() == "ND":
        lower = _to_float(_get(record, "LowerLimit")) or 0.0
        dilution = _to_float(_get(record, "Dilution"))
        dilution = dilution if dilution else 1.0
        sample_value = lower * dilution
        symbol = "<"
    else:
        sample_value = _to_float(reported)
        symbol = None

    analysis_method = _get(record, "Method")
    if mapping.method:
        analysis_method = (
            f"{analysis_method}, {mapping.method}"
            if analysis_method
            else mapping.method
        )

    analysis_date = _to_datetime(_get(record, "AnalysisTime"))
    sample_date = _to_datetime(_get(record, "SampleDate")) or analysis_date

    return {
        "analyte": mapping.analyte,
        "table": mapping.table,
        "units": str(units) if units is not None else None,
        "symbol": symbol,
        "sample_value": sample_value,
        "analysis_method": str(analysis_method) if analysis_method else None,
        "analysis_date": analysis_date,
        "sample_date": sample_date,
        "wclab_id": str(wclab_id),
        "samplepointid": str(pointid),
        "test": _get(record, "Test"),
    }


def dedupe_records(records: list[dict]) -> list[dict]:
    """Collapse duplicate (SamplePointID, WCLab_ID, Analyte) rows.

    Mirrors AMPAPI chemfile.dbprep_records: when the same analyte is reported
    more than once for the *same lab sample*, keep the ``low bromide`` test for
    Br and the ``EPA 200.7`` method for everything else. Falls back to the first
    row when no preferred method is present. Keyed on ``WCLab_ID`` too so two
    distinct lab samples for one well keep their own analyte values.
    """

    def keyf(r: dict) -> tuple[str, str, str]:
        return (r["samplepointid"], r["wclab_id"] or "", r["analyte"])

    out: list[dict] = []
    for (_pid, _wclab, analyte), group in groupby(sorted(records, key=keyf), key=keyf):
        group = list(group)
        if len(group) < 2:
            out.extend(group)
            continue

        if analyte == "Br":
            picked = next(
                (r for r in group if (r.get("test") or "").casefold() == "low bromide"),
                None,
            )
        else:
            picked = next(
                (
                    r
                    for r in group
                    if (r.get("analysis_method") or "")
                    .casefold()
                    .startswith("epa 200.7")
                ),
                None,
            )
        out.append(picked or group[0])
    return out


# --- persistence ---------------------------------------------------------------


_SUFFIX_RE_TEMPLATE = r"^{base}([A-Z]+)$"

# A PointID ending in letters is a *sample point* id, never a base well id:
# ``WL-0434A`` is a sample point on well ``WL-0434``. The base must therefore
# end in a non-letter (``WL-0434``, ``MG-030``) for the trailing letters to
# count as an incrementor.
_POINTID_SUFFIX_RE = re.compile(r"^(?P<base>.*[^A-Z])(?P<suffix>[A-Z]+)$")


def split_pointid(pointid: str) -> tuple[str, str | None]:
    """Split a PointID into its base well id and any supplied letter suffix.

    ``WL-0434A`` -> ``("WL-0434", "A")``; ``WL-0434`` -> ``("WL-0434", None)``.
    """
    match = _POINTID_SUFFIX_RE.match(pointid)
    if not match:
        return pointid, None
    return match.group("base"), match.group("suffix")


def _resolve_thing_id(session: Session, pointid: str) -> int | None:
    things = session.scalars(select(Thing).where(Thing.name == pointid)).all()
    if not things:
        return None
    # Thing.name is not guaranteed unique; take the lowest id deterministically.
    return min(t.id for t in things)


def _suffix_to_int(suffix: str) -> int:
    """Bijective base-26: A->1, B->2, ..., Z->26, AA->27, AB->28, ..."""
    n = 0
    for ch in suffix:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _int_to_suffix(n: int) -> str:
    """Inverse of :func:`_suffix_to_int` (``n`` >= 1)."""
    letters: list[str] = []
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def _existing_suffix_ints(session: Session, thing_id: int, base: str) -> set[int]:
    """Suffix numbers already used for ``base`` under this Thing.

    Chemistry sample points are the well PointID (``base``) with an appended
    letter incrementor (``A``, ``B``, ... ``Z``, ``AA``, ...). Returns the set
    of used incrementors, as bijective-base-26 integers, so the next one can be
    computed.
    """
    values = session.scalars(
        select(NMA_Chemistry_SampleInfo.nma_sample_point_id).where(
            NMA_Chemistry_SampleInfo.thing_id == thing_id
        )
    ).all()
    pattern = re.compile(_SUFFIX_RE_TEMPLATE.format(base=re.escape(base)))
    used: set[int] = set()
    for value in values:
        if not value:
            continue
        match = pattern.match(value)
        if match:
            used.add(_suffix_to_int(match.group(1)))
    return used


def _sample_exists_for_wclab(
    session: Session, thing_id: int, wclab_id: str | None
) -> bool:
    """True if this lab sample (WCLab_ID) is already recorded for the Thing."""
    # Unreachable via prep_record, which rejects a blank SampleNumber. Kept
    # because a None would compare as IS NULL and match legacy rows, silently
    # skipping a real sample.
    if wclab_id is None:
        return False
    return (
        session.scalars(
            select(NMA_Chemistry_SampleInfo.id).where(
                NMA_Chemistry_SampleInfo.thing_id == thing_id,
                NMA_Chemistry_SampleInfo.nma_wclab_id == wclab_id,
            )
        ).first()
        is not None
    )


def _build_measurement(
    model, chemistry_sample_info_id: int, rec: dict, sample_point_id: str
):
    analysis_date = rec["analysis_date"]
    if model is NMA_MinorTraceChemistry and isinstance(analysis_date, datetime):
        # NMA_MinorTraceChemistry.analysis_date is a DATE column.
        analysis_date = analysis_date.date()
    return model(
        chemistry_sample_info_id=chemistry_sample_info_id,
        nma_global_id=uuid.uuid4(),
        nma_sample_point_id=sample_point_id,
        nma_wclab_id=rec["wclab_id"],
        analyte=rec["analyte"],
        symbol=rec["symbol"],
        sample_value=rec["sample_value"],
        units=rec["units"],
        analysis_method=rec["analysis_method"],
        analysis_date=analysis_date,
        analyses_agency=ANALYSES_AGENCY,
    )


def bulk_upload_chemistry(
    source: Path | str | bytes,
) -> ChemistryUploadResult:
    """Ingest a LIMS ``.xlsx`` workbook into the NMA chemistry tables.

    ``source`` may be a filesystem path or the raw ``.xlsx`` bytes (e.g. a file
    downloaded from Google Drive).

    The workbook's ``SamplePointID`` is the base well PointID. Each distinct lab
    sample (``WCLab_ID`` / SampleNumber) for a well becomes a new
    ``NMA_Chemistry_SampleInfo`` row whose ``nma_sample_point_id`` is the base
    with the next letter incrementor appended (``A``, ``B``, ... ``Z``, ``AA``,
    ...). A lab sample already recorded for the well (same ``WCLab_ID``) is
    skipped, so re-running is idempotent.

    A data-quality problem aborts the whole file and nothing is written: a row
    that fails to map, a row with no ``SampleNumber`` (the WCLab_ID that makes a
    re-ingest recognizable), or a ``SamplePointID`` with no matching Thing.
    """
    if isinstance(source, str):
        source = Path(source)

    try:
        raw_records = read_lims_xlsx(source)
    except Exception as exc:  # openpyxl raises a variety of parse errors
        return _result(
            processed=0,
            imported=0,
            validation_errors=[f"Could not read workbook: {exc}"],
            skipped_duplicates=[],
            created=[],
        )

    processed = len(raw_records)
    validation_errors: list[str] = []
    prepped: list[dict] = []
    for offset, raw in enumerate(raw_records):
        # +2: worksheet row 1 is the header, enumerate is 0-based.
        row_number = offset + 2
        try:
            prepped.append(prep_record(raw))
        except ChemistryMappingError as exc:
            validation_errors.append(f"Row {row_number}: {exc}")

    prepped = dedupe_records(prepped)

    warnings: list[str] = []

    with session_ctx() as session:
        # A workbook's SamplePointID may already carry a sample-point letter
        # (WL-0434A). The well is always the base (WL-0434), so strip it before
        # resolving; the supplied letter is checked against the computed one
        # below.
        supplied_suffixes: dict[str, str | None] = {}
        for rec in prepped:
            base, suffix = split_pointid(rec["samplepointid"])
            rec["samplepointid"] = base
            # Keep the first supplied suffix seen for the well; a workbook
            # should not disagree with itself, and if it does the mismatch
            # warning below still fires.
            supplied_suffixes.setdefault(base, suffix)

        # Resolve every distinct (base) sample point to a Thing up front.
        base_pointids = sorted({r["samplepointid"] for r in prepped})
        thing_ids: dict[str, int | None] = {
            pid: _resolve_thing_id(session, pid) for pid in base_pointids
        }
        for pid in base_pointids:
            if thing_ids[pid] is None:
                validation_errors.append(
                    f"SamplePointID {pid}: no matching Thing (well) found"
                )

        # Abort the whole file on any data-quality problem before writing.
        if validation_errors:
            return _result(
                processed=processed,
                imported=0,
                validation_errors=validation_errors,
                skipped_duplicates=[],
                created=[],
            )

        # One sample = one lab sample (WCLab_ID) for a well.
        def bucket_key(r: dict) -> tuple[str, str | None]:
            return (r["samplepointid"], r["wclab_id"])

        buckets: dict[tuple[str, str | None], list[dict]] = {}
        for rec in sorted(
            prepped, key=lambda r: (r["samplepointid"], r["wclab_id"] or "")
        ):
            buckets.setdefault(bucket_key(rec), []).append(rec)

        # Per-Thing set of used suffix incrementors, seeded from the DB and
        # extended as we assign new ones within this run.
        used_suffixes: dict[int, set[int]] = {}
        skipped_duplicates: list[dict] = []
        created: list[dict] = []
        imported = 0

        for (base, wclab_id), recs in buckets.items():
            thing_id = thing_ids[base]

            # Already ingested this lab sample -> skip (idempotent), keep going.
            if _sample_exists_for_wclab(session, thing_id, wclab_id):
                skipped_duplicates.append({"pointid": base, "wclab_id": wclab_id})
                continue

            if thing_id not in used_suffixes:
                used_suffixes[thing_id] = _existing_suffix_ints(session, thing_id, base)
            next_int = (
                max(used_suffixes[thing_id]) + 1 if used_suffixes[thing_id] else 1
            )
            used_suffixes[thing_id].add(next_int)
            computed_suffix = _int_to_suffix(next_int)
            sample_point_id = f"{base}{computed_suffix}"

            # The workbook may have supplied its own letter. The computed one
            # wins (it cannot collide with an existing sample point), but a
            # disagreement is surfaced so a human can reconcile it.
            supplied = supplied_suffixes.get(base)
            if supplied is not None and supplied != computed_suffix:
                warnings.append(
                    f"{base}: workbook supplied sample point {base}{supplied}, "
                    f"but the next free incrementor is {computed_suffix}; "
                    f"loaded as {sample_point_id}."
                )

            collection_date = next(
                (r["sample_date"] for r in recs if r["sample_date"]), None
            )
            info = NMA_Chemistry_SampleInfo(
                thing_id=thing_id,
                nma_sample_pt_id=uuid.uuid4(),
                nma_sample_point_id=sample_point_id,
                nma_wclab_id=wclab_id,
                analyses_agency=ANALYSES_AGENCY,
                collection_date=collection_date,
            )
            session.add(info)
            session.flush()  # assign info.id for the FK below

            for rec in recs:
                model = _TABLE_MODEL[rec["table"]]
                session.add(_build_measurement(model, info.id, rec, sample_point_id))
                imported += 1
            created.append(
                {
                    "sample_point_id": sample_point_id,
                    "wclab_id": wclab_id,
                    "rows": len(recs),
                }
            )

        session.commit()

    return _result(
        processed=processed,
        imported=imported,
        validation_errors=validation_errors,
        skipped_duplicates=skipped_duplicates,
        created=created,
        warnings=warnings,
    )


def _result(
    *,
    processed: int,
    imported: int,
    validation_errors: list[str],
    skipped_duplicates: list[dict],
    created: list[dict],
    warnings: list[str] | None = None,
) -> ChemistryUploadResult:
    warnings = warnings or []
    rows_with_issues = len(validation_errors) + len(skipped_duplicates) + len(warnings)
    payload = {
        "summary": {
            "total_rows_processed": processed,
            "total_rows_imported": imported,
            "validation_errors_or_warnings": rows_with_issues,
            "samples_created": len(created),
            "samples_skipped": len(skipped_duplicates),
        },
        "validation_errors": validation_errors,
        "warnings": warnings,
        "skipped_duplicates": skipped_duplicates,
        "created_samples": created,
    }
    stderr_parts: list[str] = []
    if validation_errors:
        stderr_parts.append("\n".join(validation_errors))
    if warnings:
        stderr_parts.append("\n".join(warnings))
    if skipped_duplicates:
        dupes = ", ".join(
            f"{d['pointid']} (WCLab_ID {d['wclab_id']})" for d in skipped_duplicates
        )
        stderr_parts.append(f"Skipped already-ingested lab sample(s): {dupes}")
    stderr = "\n".join(stderr_parts)
    # Only a data-quality abort is a failure; skipped duplicates are idempotent.
    exit_code = 1 if validation_errors else 0
    return ChemistryUploadResult(exit_code=exit_code, stderr=stderr, payload=payload)


# ============= EOF =============================================
