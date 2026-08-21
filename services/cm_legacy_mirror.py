# ===============================================================================
# Copyright 2026 ross
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ===============================================================================
"""Load the McLemore critical-minerals workbook into the CM_legacy mirror.

Reads ``McLemoreMasterChem*.xlsx`` and writes each sheet, unchanged, into the
``CM_*`` staging tables (see ``db/cm_legacy.py``). No interpretation happens
here: every cell becomes text, censored values keep their ``<`` prefix, Excel
error text is carried through, and the ChemicalData / GIS / QAQC sheets are all
loaded even though they disagree with each other.

The load is idempotent per sheet -- existing rows for a sheet are deleted before
its rows are inserted -- so a revised workbook can be reloaded without
duplicating rows or resetting the whole mirror.

Layout is asserted, not guessed. Each sheet declares which 1-based row holds
its header, and the load fails if that row does not contain the expected label.
A revised workbook that moves a header row, renames a column, or adds one must
be looked at by a human before it lands in the mirror.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterator, Sequence

from sqlalchemy import delete, insert
from sqlalchemy.orm import Session

from db.cm_legacy import (
    CM_SHEET_CHEMICAL_DATA,
    CM_SHEET_GIS,
    CM_SHEET_QAQC,
    SOURCE_HEADER_BY_COLUMN,
    CM_ChemicalData,
    CM_DetectionLimits,
    CM_MineralSystems,
    CM_References,
    CM_WorkbookMetadata,
    CM_WorldComparisons,
    CM_WorldReferences,
)

# 1-based row holding the column headers, and the first row of data, for each
# sheet mirrored into CM_ChemicalData. ChemicalData carries a title banner in
# row 1 and a units row in row 3; GIS is the same columns with neither.
CHEMISTRY_SHEET_LAYOUT: dict[str, tuple[int, int]] = {
    CM_SHEET_CHEMICAL_DATA: (2, 4),
    CM_SHEET_GIS: (1, 2),
    CM_SHEET_QAQC: (2, 3),
}

# Sheets flattened into CM_WorkbookMetadata as (label, value) pairs.
METADATA_SHEETS = ("General Information", "MetaData", "DefinitionOfFields")

MINERAL_SYSTEM_COLUMNS = (
    "system_name",
    "synopsis",
    "deposit_types",
    "principal_commodities",
    "critical_minerals",
    "references",
    "phase_2",
    "phase_3",
    "phase_4",
)

INSERT_BATCH_SIZE = 1000


class CMWorkbookError(ValueError):
    """The workbook does not have the layout the mirror was built for."""


@dataclass
class CMMirrorLoadResult:
    """Rows written per mirror table, plus anything the loader wants flagged."""

    rows_by_sheet: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return sum(self.rows_by_sheet.values())


def normalize_header(value: Any) -> str:
    return " ".join(str(value).split()).lower()


def cell_to_text(value: Any) -> str | None:
    """Render a cell as text without losing or reinterpreting anything.

    Dates become ISO-8601 (the workbook stores them as Excel serials, which
    would otherwise mirror as meaningless integers), numbers keep Python's
    round-trippable repr, and everything else -- including ``<0.1`` and
    ``#VALUE!`` -- is passed through with surrounding whitespace stripped.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return (
            value.isoformat(sep=" ")
            if any((value.hour, value.minute, value.second, value.microsecond))
            else value.date().isoformat()
        )
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    text = str(value).strip()
    return text or None


def _is_blank(row: Sequence[Any]) -> bool:
    return all(cell_to_text(v) is None for v in row)


def column_by_header(table_name: str) -> dict[str, str]:
    """Invert SOURCE_HEADER_BY_COLUMN into {normalized header: column}."""
    return {
        normalize_header(header): column
        for column, header in SOURCE_HEADER_BY_COLUMN[table_name].items()
    }


def _map_headers(
    sheet_name: str,
    header_row: Sequence[Any],
    columns_by_header: dict[str, str],
) -> dict[int, str]:
    """Map each populated header cell to its mirror column, or fail."""
    mapping: dict[int, str] = {}
    unknown: list[str] = []
    for index, cell in enumerate(header_row):
        if cell_to_text(cell) is None:
            continue
        column = columns_by_header.get(normalize_header(cell))
        if column is None:
            unknown.append(str(cell).strip())
            continue
        mapping[index] = column
    if unknown:
        raise CMWorkbookError(
            f"sheet {sheet_name!r} has {len(unknown)} header(s) the mirror has no "
            f"column for: {unknown[:10]}. The workbook changed shape; update "
            f"db/cm_legacy.py (and a migration) before loading it."
        )
    if not mapping:
        raise CMWorkbookError(f"sheet {sheet_name!r} has no recognizable headers")
    return mapping


def _rows(worksheet: Any) -> list[tuple[Any, ...]]:
    return list(worksheet.iter_rows(values_only=True))


def _iter_sheet_records(
    worksheet: Any,
    sheet_name: str,
    header_row_number: int,
    first_data_row_number: int,
    columns_by_header: dict[str, str],
    expected_header: str,
) -> Iterator[dict[str, Any]]:
    rows = _rows(worksheet)
    if len(rows) < first_data_row_number:
        raise CMWorkbookError(f"sheet {sheet_name!r} has no data rows")

    header_row = rows[header_row_number - 1]
    if expected_header not in {
        normalize_header(c) for c in header_row if c is not None
    }:
        raise CMWorkbookError(
            f"sheet {sheet_name!r} row {header_row_number} is not the header row "
            f"(expected to find {expected_header!r})"
        )

    mapping = _map_headers(sheet_name, header_row, columns_by_header)
    for offset, row in enumerate(rows[first_data_row_number - 1 :]):
        if _is_blank(row):
            continue
        record: dict[str, Any] = {}
        for index, column in mapping.items():
            if index < len(row):
                record[column] = cell_to_text(row[index])
        record["source_row"] = first_data_row_number + offset
        yield record


def _replace(
    session: Session, model: Any, records: list[dict[str, Any]], **scope: Any
) -> int:
    """Delete the scoped rows, then insert the given ones."""
    statement = delete(model)
    for column, value in scope.items():
        statement = statement.where(getattr(model, column) == value)
    session.execute(statement)
    for start in range(0, len(records), INSERT_BATCH_SIZE):
        batch = records[start : start + INSERT_BATCH_SIZE]
        if batch:
            session.execute(insert(model), batch)
    return len(records)


def _load_chemistry_sheets(workbook: Any, session: Session) -> dict[str, int]:
    columns_by_header = column_by_header("CM_ChemicalData")
    counts: dict[str, int] = {}
    for sheet_name, (header_row, first_data_row) in CHEMISTRY_SHEET_LAYOUT.items():
        if sheet_name not in workbook.sheetnames:
            raise CMWorkbookError(f"workbook is missing the {sheet_name!r} sheet")
        records = [
            dict(record, source_sheet=sheet_name)
            for record in _iter_sheet_records(
                workbook[sheet_name],
                sheet_name,
                header_row,
                first_data_row,
                columns_by_header,
                expected_header="sample",
            )
        ]
        counts[sheet_name] = _replace(
            session, CM_ChemicalData, records, source_sheet=sheet_name
        )
    return counts


def _load_detection_limits(workbook: Any, session: Session) -> int:
    rows = _rows(workbook["DetectionLimits"])
    method = cell_to_text(rows[0][0]) if rows else None
    records = [
        {
            "source_row": number,
            "method": method,
            "element": cell_to_text(row[0]),
            "lower_reporting_limit": cell_to_text(row[1]) if len(row) > 1 else None,
            "unit": cell_to_text(row[2]) if len(row) > 2 else None,
        }
        for number, row in enumerate(rows[2:], start=3)
        if not _is_blank(row)
    ]
    return _replace(session, CM_DetectionLimits, records)


def _load_citations(workbook: Any, session: Session, sheet: str, model: Any) -> int:
    records = [
        {"source_row": number, "citation": cell_to_text(row[0])}
        for number, row in enumerate(_rows(workbook[sheet]), start=1)
        if not _is_blank(row)
    ]
    return _replace(session, model, records)


def _load_mineral_systems(workbook: Any, session: Session) -> int:
    records = []
    for number, row in enumerate(_rows(workbook["MineralSystems"]), start=1):
        if _is_blank(row):
            continue
        record: dict[str, Any] = {"source_row": number}
        for index, column in enumerate(MINERAL_SYSTEM_COLUMNS):
            record[column] = cell_to_text(row[index]) if index < len(row) else None
        records.append(record)
    return _replace(session, CM_MineralSystems, records)


def _load_world_comparisons(workbook: Any, session: Session) -> int:
    records = list(
        _iter_sheet_records(
            workbook["world"],
            "world",
            header_row_number=1,
            first_data_row_number=2,
            columns_by_header=column_by_header("CM_WorldComparisons"),
            expected_header="area",
        )
    )
    return _replace(session, CM_WorldComparisons, records)


def _load_workbook_metadata(workbook: Any, session: Session) -> dict[str, int]:
    counts: dict[str, int] = {}
    for sheet in METADATA_SHEETS:
        records = [
            {
                "source_sheet": sheet,
                "source_row": number,
                "label": cell_to_text(row[0]) if row else None,
                "value": cell_to_text(row[1]) if len(row) > 1 else None,
            }
            for number, row in enumerate(_rows(workbook[sheet]), start=1)
            if not _is_blank(row)
        ]
        counts[sheet] = _replace(
            session, CM_WorkbookMetadata, records, source_sheet=sheet
        )
    return counts


def load_cm_workbook(source_file: Path | str, session: Session) -> CMMirrorLoadResult:
    """Mirror every sheet of the critical-minerals workbook into CM_* tables.

    The caller owns the transaction, so a failure part way through leaves the
    mirror untouched rather than half-loaded.
    """
    import openpyxl

    source_file = Path(source_file)
    if not source_file.exists():
        raise CMWorkbookError(f"workbook not found: {source_file}")

    workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=True)

    required_sheets = {
        "DetectionLimits",
        "References",
        "MineralSystems",
        "world",
        "world_ref",
        *METADATA_SHEETS,
    }
    missing = sorted(required_sheets - set(workbook.sheetnames))
    if missing:
        workbook.close()
        raise CMWorkbookError(f"workbook is missing required sheet(s): {missing}")

    try:
        result = CMMirrorLoadResult()
        result.rows_by_sheet.update(_load_chemistry_sheets(workbook, session))
        result.rows_by_sheet["DetectionLimits"] = _load_detection_limits(workbook, session)
        result.rows_by_sheet["References"] = _load_citations(
            workbook, session, "References", CM_References
        )
        result.rows_by_sheet["MineralSystems"] = _load_mineral_systems(
            workbook, session
        )
        result.rows_by_sheet["world"] = _load_world_comparisons(workbook, session)
        result.rows_by_sheet["world_ref"] = _load_citations(
            workbook, session, "world_ref", CM_WorldReferences
        )
        result.rows_by_sheet.update(_load_workbook_metadata(workbook, session))
    finally:
        workbook.close()

    chemical_data = result.rows_by_sheet.get(CM_SHEET_CHEMICAL_DATA, 0)
    gis = result.rows_by_sheet.get(CM_SHEET_GIS, 0)
    if chemical_data != gis:
        result.warnings.append(
            f"ChemicalData ({chemical_data} rows) and GIS ({gis} rows) disagree on row "
            f"count; the sheets are known to have drifted apart and reconciliation is "
            f"deferred (see db/cm_legacy.py)"
        )
    return result
